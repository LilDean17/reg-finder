#!/usr/bin/env python3
"""从 output/results.jsonl 提取 score > 5 的结果，生成 xlsx"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUT = Path("output/results.jsonl")
OUTPUT = Path("output/results_filtered.xlsx")
THRESHOLD = 5

# 读取并过滤
results = []
with open(INPUT, "r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        r = json.loads(line)
        if r["score"] > THRESHOLD:
            results.append(r)

results.sort(key=lambda x: -x["score"])

# 构建 breakdown 字符串
def fmt_breakdown(r):
    parts = []
    for d in r["breakdown"]:
        if d["weight"] != 0:
            w = f"+{d['weight']}" if d["weight"] > 0 else str(d["weight"])
            parts.append(f"{w} [{d['profile']}] {d['indicator']}")
    return "; ".join(parts)

# 创建 xlsx
wb = Workbook()
ws = wb.active
ws.title = f"得分>{THRESHOLD}的结果"

# 列头
headers = ["得分", "URL", "最终URL", "状态码", "标题", "系统类型",
           "SPA", "已渲染", "注册表单", "推荐等级", "命中规则明细", "错误"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 颜色填充
fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# 数据行
for row_idx, r in enumerate(results, 2):
    biz = ",".join(r["business_types"]) if r["business_types"] else ""
    bd = fmt_breakdown(r)
    title = (r["title"] or "")[:80]

    row_data = [
        r["score"],
        r["url"],
        r["final_url"],
        r["status_code"],
        title,
        biz,
        "是" if r["is_spa"] else "否",
        "是" if r["rendered"] else "否",
        "是" if r["has_register_form"] else "否",
        r["recommendation"],
        bd,
        r["error"],
    ]

    score = r["score"]
    if score >= 80:
        row_fill = fill_red
    elif score >= 60:
        row_fill = fill_yellow
    else:
        row_fill = fill_green

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# 调整列宽
col_widths = [6, 40, 40, 8, 30, 15, 6, 8, 10, 10, 60, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

# 冻结首行
ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"[*] 已保存 {len(results)} 条结果到 {OUTPUT}")
print(f"    得分范围: {results[-1]['score']} ~ {results[0]['score']}")
