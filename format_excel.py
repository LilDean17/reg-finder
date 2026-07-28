
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

dir_path = r"D:\工具\脆弱性打分\reg-finder\output"
xlsx_path = f"{dir_path}\\results_filtered.xlsx"

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# --- Styles ---
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_font = Font(size=11)
cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# --- Column widths ---
col_widths = {
    "A": 12,   # 得分
    "B": 45,   # URL
    "C": 45,   # 最终URL
    "D": 10,   # 状态码
    "E": 35,   # 标题
    "F": 25,   # 业务类型
    "G": 8,    # SPA
    "H": 10,   # 已渲染
    "I": 12,   # 注册表单
    "J": 14,   # 推荐等级
    "K": 80,   # 命中规则明细
    "L": 40,   # 错误
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# --- Header row formatting ---
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# --- Data rows formatting ---
# Auto row height based on content length
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    max_lines = 1
    for cell in row:
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = center_align if cell.column <= 10 else cell_align

        # Estimate line count for wrap_text
        if cell.value:
            text_len = len(str(cell.value))
            if cell.column == 1:   # 得分
                max_lines = 1
            elif cell.column in (11, 12):  # 命中规则明细, 错误
                lines = max(1, text_len // 55)
                max_lines = max(max_lines, lines)
            elif cell.column in (7, 8, 9):  # SPA, 已渲染, 注册表单
                max_lines = 1
            else:
                lines = max(1, text_len // 30)
                max_lines = max(max_lines, lines)

    # Set row height (min 20, max 200)
    row_height = max(20, min(200, max_lines * 20))
    ws.row_dimensions[row_idx].height = row_height

# --- Freeze header row ---
ws.freeze_panes = "A2"

# --- Auto filter ---
ws.auto_filter.ref = ws.dimensions

wb.save(xlsx_path)
print(f"Formatted and saved: {xlsx_path}")
print(f"Total data rows: {ws.max_row - 1}")
