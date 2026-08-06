#!/usr/bin/env python3
"""输出模块：终端彩色表格 + JSON + CSV + XLSX"""
import json
import csv
import os
from dataclasses import asdict
from pathlib import Path
from datetime import datetime

import colorama
from colorama import Fore, Style
colorama.init()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from core.config import OutputConfig


class OutputFormatter:
    _RT_SHORT = {
        "body_keyword": "body",
        "register_form": "form",
        "profile_base": "base",
        "path_keyword": "path",
        "title_keyword": "title",
        "domain_keyword": "domain",
        "status_code": "status",
        "spa_rendered": "spa",
    }

    def __init__(self, output_config: OutputConfig):
        self.config = output_config
        Path("output").mkdir(exist_ok=True)

    @classmethod
    def _rt(cls, rule_type: str) -> str:
        return cls._RT_SHORT.get(rule_type, rule_type)

    def print(self, results: list, total_urls: int, elapsed: float):
        """根据配置的输出格式输出结果"""
        if "terminal" in self.config.formats:
            self._print_terminal(results, total_urls, elapsed)

        if "json" in self.config.formats:
            self._save_json(results)

        if "csv" in self.config.formats:
            self._save_csv(results)

        if "xlsx" in self.config.formats:
            self._save_xlsx(results)

    def _print_terminal(self, results: list, total_urls: int, elapsed: float):
        """终端彩色表格输出"""
        if not results:
            print(f"\n[!] 未找到高于 {self.config.min_score} 分的结果")
            return

        # 计算列宽
        url_width = max(len(r.final_url) for r in results)
        url_width = min(url_width, 55)

        divider = "─" * (8 + url_width + 10 + 14 + 4 + 5 + 6 + 10)

        print(f"\n{'═' * len(divider)}")
        print(f"  注册可行性筛选结果  (共 {len(results)} 个 / 扫描 {total_urls} 个 / 耗时 {elapsed:.1f}s)")
        print(f"{'═' * len(divider)}")

        header = f"{'得分':<6} {'URL':<{url_width}} {'状态':<6} {'业务类型':<14} {'SPA':<4} {'渲染':<5} {'渲染状态':<6} {'注册表单'}"
        print(header)
        print(divider)

        for r in results:
            score_str = self._color_score(r.score)
            url_short = r.final_url[:url_width]
            biz = ",".join(r.business_types) if r.business_types else "未分类"
            biz = biz[:12]
            spa = "是" if r.is_spa else ""
            rendered = "✓" if r.rendered else ""
            render_status = "已渲染" if r.rendered else ("未渲染" if r.is_spa else "-")
            form = "✓" if r.has_register_form else ""

            print(
                f"{score_str:<6} {url_short:<{url_width}} "
                f"{r.status_code:<6} {biz:<14} {spa:<4} {rendered:<5} {render_status:<6} {form}"
            )

            # 输出每条命中规则
            for d in r.breakdown:
                weight_str = f"+{d.weight}" if d.weight > 0 else f"{d.weight}"
                print(f"       │  {weight_str:>4}  [{d.profile}][{self._rt(d.rule_type)}] {d.indicator}")

            print(divider)

    def _color_score(self, score: int) -> str:
        """分数着色"""
        if score >= 80:
            return f"{Fore.RED}{score}{Style.RESET_ALL}"    # 红色：高价值
        elif score >= 60:
            return f"{Fore.YELLOW}{score}{Style.RESET_ALL}"  # 黄色：过线
        else:
            return f"{Fore.GREEN}{score}{Style.RESET_ALL}"   # 绿色：低分

    def _save_json(self, results: list):
        path = Path("output/results.json")
        data = [self._serialize(r) for r in results]
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        print(f"\n[*] JSON 已保存: {path}")

    def _save_csv(self, results: list):
        path = Path("output/results.csv")
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "得分", "URL", "最终URL", "状态码", "标题",
                "业务类型", "SPA", "已渲染", "渲染状态", "注册表单", "推荐等级",
                "命中规则明细", "错误"
            ])
            for r in results:
                breakdown_str = "; ".join(
                    f"{'+' if d.weight > 0 else ''}{d.weight} [{d.profile}][{OutputFormatter._rt(d.rule_type)}] {d.indicator}"
                    for d in r.breakdown
                )
                writer.writerow([
                    r.score,
                    r.url,
                    r.final_url,
                    r.status_code,
                    r.title[:80] if r.title else "",
                    ",".join(r.business_types) if r.business_types else "",
                    "是" if r.is_spa else "否",
                    "是" if r.rendered else "否",
                    "已渲染" if r.rendered else ("未渲染" if r.is_spa else "-"),
                    "是" if r.has_register_form else "否",
                    r.recommendation,
                    breakdown_str,
                    r.error,
                ])
        print(f"[*] CSV 已保存: {path}")

    def _save_xlsx(self, results: list):
        """保存全部结果到 XLSX，同时导出 score > 5 的高分结果到单独文件"""
        if not HAS_OPENPYXL:
            print("\n[!] 未安装 openpyxl，跳过 XLSX 输出（pip install openpyxl）")
            return

        # 全部结果
        all_path = Path("output/results.xlsx")
        self._write_xlsx(all_path, results)
        print(f"\n[*] XLSX 已保存: {all_path}")

        # 高分筛选（score > 5）
        high = [r for r in results if r.score > 5]
        if high:
            high_path = Path("output/high_score_above5.xlsx")
            self._write_xlsx(high_path, high)
            print(f"[*] 高分结果 (score>5) XLSX 已保存: {high_path}  ({len(high)} 条)")
        else:
            print(f"\n[*] 无 score > 5 的结果，未生成高分 XLSX")

    def save_all_xlsx(self, results: list):
        """仅保存全部结果的 XLSX（高分文件由流式追加单独处理）"""
        if not HAS_OPENPYXL:
            return
        all_path = Path("output/results.xlsx")
        self._write_xlsx(all_path, results)
        print(f"\n[*] XLSX 已保存: {all_path}")

    @staticmethod
    def _has_register_signal(result) -> bool:
        """判断结果是否命中注册信号：score < 5 且命中规则明细中包含注册/register相关字眼"""
        if result.score >= 5:
            return False

        # 注册相关关键词（含变体：去空格、去分隔符）
        keywords = [
            # 中文
            "注册", "注 册", "注-册", "注.册", "註冊",
            "报 名", "报-名", "登 记", "登-记",
            "开 户", "开-户", "申 请", "申-请",
            "马上注册", "立即注册", "免费注册",
            # 英文
            "register", "regist", "sign up", "signup", "sign-up",
            "create account", "new user", "join us",
            "become a member", "subscribe",
        ]

        for d in result.breakdown:
            indicator_lower = d.indicator.lower()
            if any(kw.lower() in indicator_lower for kw in keywords):
                return True

        # 也检查表单检测标志
        if result.has_register_form:
            return True
        return False

    def save_register_signal_xlsx(self, results: list):
        """将 score<5 但命中注册信号的结果单独导出到 XLSX"""
        if not HAS_OPENPYXL:
            print("\n[!] 未安装 openpyxl，跳过注册信号 XLSX 输出")
            return

        filtered = [r for r in results if self._has_register_signal(r)]
        if not filtered:
            print(f"\n[*] 无 score<5 且命中注册信号的结果，未生成注册信号 XLSX")
            return

        path = Path("output/register_signal_low_score.xlsx")
        self._write_xlsx(path, filtered)
        print(f"[*] 注册信号结果 (score<5+注册命中) XLSX 已保存: {path}  ({len(filtered)} 条)")

    # ── 实时流式 XLSX（扫描过程中逐条追加） ──────────────────────────

    def create_streaming_xlsx(self, path: Path) -> "Workbook":
        """创建一个带表头的工作簿，供扫描过程中逐条追加使用"""
        wb = Workbook()
        ws = wb.active
        ws.title = "高分红分"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

        headers = [
            "得分", "URL", "最终URL", "状态码", "标题",
            "业务类型", "SPA", "已渲染", "渲染状态", "注册表单", "推荐等级",
            "命中规则明细", "错误"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        col_widths = [8, 40, 40, 10, 40, 20, 6, 8, 10, 10, 14, 50, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.freeze_panes = "A2"
        path.parent.mkdir(exist_ok=True)
        wb._output_path = str(path)  # type: ignore[attr-defined]
        wb.save(str(path))
        return wb

    def append_streaming_row(self, wb: "Workbook", result) -> None:
        """向流式 XLSX 追加一行并保存到磁盘"""
        ws = wb.active
        breakdown_str = "; ".join(
            f"{'+' if d.weight > 0 else ''}{d.weight} [{d.profile}][{d.rule_type[:2]}] {d.indicator}"
            for d in result.breakdown
        )
        row = [
            result.score,
            result.url,
            result.final_url,
            result.status_code,
            result.title[:80] if result.title else "",
            ",".join(result.business_types) if result.business_types else "",
            "是" if result.is_spa else "否",
            "是" if result.rendered else "否",
            "已渲染" if result.rendered else ("未渲染" if result.is_spa else "-"),
            "是" if result.has_register_form else "否",
            result.recommendation,
            breakdown_str,
            result.error,
        ]
        ws.append(row)
        row_idx = ws.max_row

        # 分数列着色
        score_cell = ws.cell(row=row_idx, column=1)
        if result.score >= 80:
            score_cell.font = Font(color="FF0000", bold=True)
        elif result.score >= 60:
            score_cell.font = Font(color="FFC000", bold=True)
        else:
            score_cell.font = Font(color="00B050", bold=True)
        score_cell.alignment = Alignment(horizontal="center", vertical="center")
        score_cell.border = Border(
            left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
        )

        wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col_idx in range(2, len(row) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = Border(
                left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
            )
            cell.alignment = wrap if col_idx in (12, 13) else Alignment(horizontal="left", vertical="center")

        wb.save(wb._output_path)  # type: ignore[attr-defined]

    def _write_xlsx(self, path: Path, results: list):
        """通用 XLSX 写入"""
        wb = Workbook()
        ws = wb.active
        ws.title = "扫描结果"

        # 样式定义
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
        wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 表头
        headers = [
            "得分", "URL", "最终URL", "状态码", "标题",
            "业务类型", "SPA", "已渲染", "渲染状态", "注册表单", "推荐等级",
            "命中规则明细", "错误"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 分数着色条件
        def _score_color(score: int) -> str:
            if score >= 80:
                return "FF0000"
            elif score >= 60:
                return "FFC000"
            return "00B050"

        # 数据行
        for r in results:
            breakdown_str = "; ".join(
                f"{'+' if d.weight > 0 else ''}{d.weight} [{d.profile}][{d.rule_type[:2]}] {d.indicator}"
                for d in r.breakdown
            )
            row = [
                r.score,
                r.url,
                r.final_url,
                r.status_code,
                r.title[:80] if r.title else "",
                ",".join(r.business_types) if r.business_types else "",
                "是" if r.is_spa else "否",
                "是" if r.rendered else "否",
                "已渲染" if r.rendered else ("未渲染" if r.is_spa else "-"),
                "是" if r.has_register_form else "否",
                r.recommendation,
                breakdown_str,
                r.error,
            ]
            ws.append(row)
            row_idx = ws.max_row

            # 分数列着色
            score_cell = ws.cell(row=row_idx, column=1)
            score_cell.font = Font(color=_score_color(r.score), bold=True)
            score_cell.alignment = center_align
            score_cell.border = thin_border

            # 其余列样式
            for col_idx in range(2, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = wrap_align if col_idx in (12, 13) else left_align

        # 自动列宽
        col_widths = [8, 40, 40, 10, 40, 20, 6, 8, 10, 10, 14, 50, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        path.parent.mkdir(exist_ok=True)
        wb.save(str(path))

    def _serialize(self, r) -> dict:
        return {
            "score": r.score,
            "url": r.url,
            "final_url": r.final_url,
            "status_code": r.status_code,
            "title": r.title,
            "business_types": r.business_types,
            "is_spa": r.is_spa,
            "rendered": r.rendered,
            "has_register_form": r.has_register_form,
            "recommendation": r.recommendation,
            "breakdown": [
                {
                    "profile": d.profile,
                    "category": d.category,
                    "rule_type": d.rule_type,
                    "indicator": d.indicator,
                    "weight": d.weight,
                }
                for d in r.breakdown
            ],
            "error": r.error,
        }
